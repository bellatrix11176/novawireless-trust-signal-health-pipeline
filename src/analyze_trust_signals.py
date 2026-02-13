#!/usr/bin/env python3
"""
analyze_trust_signals.py  (ONE-SCRIPT EDITION + CRT)

Single source of truth:
- Reads ONE input file from /data (default: novawireless_synthetic_calls.csv)
- Writes ALL derived artifacts to /output

Includes:
- Optional PII scrubbing (hash customer_phone, optional interaction_id; drop PII columns + free-text fields)
- Feature engineering (repeat contact + dispositions)
- Drift score + summaries (rep / ivr / customer repeats)
- CRT support:
  - Detects CRT column (CRT_seconds, CRT, crt, etc.)
  - Computes CRT risk score (0-1) based on thresholds
  - Adds CRT aggregates to rep/ivr summaries
  - Adds CRT to correlation heatmaps
  - Adds CRT section to summary_report.txt

Outputs:
- output/calls_scrubbed.csv (optional; safe for GitHub)
- output/calls_scored_cleaned.csv (scored dataset)
- output/rep_summary.csv
- output/ivr_summary.csv
- output/customer_repeat_summary.csv
- output/summary_report.txt
- output/*.png charts + story charts + correlation heatmaps
- output/Trust_Signal_Health_Psuedo_Math_Check.xlsx (optional mathcheck)

Usage:
  python src/analyze_trust_signals.py run
  python src/analyze_trust_signals.py run --with_mathcheck
  python src/analyze_trust_signals.py run --file novawireless_synthetic_calls.csv
  python src/analyze_trust_signals.py run --crt_low 480 --crt_high 900
  python src/analyze_trust_signals.py scrub
  python src/analyze_trust_signals.py charts
  python src/analyze_trust_signals.py mathcheck
"""

from __future__ import annotations

import argparse
import hashlib
import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------
# Repo path helpers (repo-root style)
# ---------------------------

def find_repo_root(start: Path) -> Path:
    """
    Finds the folder that contains:
      - data/
      - src/
    Ensures output/ exists (creates if missing).
    """
    cur = start.resolve()
    for _ in range(50):
        if (cur / "data").is_dir() and (cur / "src").is_dir():
            (cur / "output").mkdir(parents=True, exist_ok=True)
            return cur
        if cur.parent == cur:
            break
        cur = cur.parent
    raise FileNotFoundError(
        "Could not find repo root containing data/ and src/.\n"
        "Run this script from somewhere inside your trust_signals folder."
    )


def get_input_file(data_dir: Path, filename: str) -> Path:
    p = data_dir / filename
    if not p.exists():
        raise FileNotFoundError(
            f"Expected input file not found: {p}\n"
            "Place your dataset in /data and rerun."
        )
    return p


def load_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    return df


# ---------------------------
# PII scrubbing
# ---------------------------

def _hash_series(series: pd.Series, salt: str) -> pd.Series:
    def h(v: object) -> str:
        s = "" if pd.isna(v) else str(v).strip()
        raw = (salt + "|" + s).encode("utf-8")
        return hashlib.sha256(raw).hexdigest()
    return series.apply(h)


def scrub_pii_df(
    df: pd.DataFrame,
    salt: str = "trust-signals-salt",
    hash_interaction_id: bool = True,
    keep_text: bool = False,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Returns:
      scrubbed_df, dropped_columns

    Required columns for downstream analysis: timestamp, customer_phone
    Hashes customer_phone to stable token so repeat-contact logic still works.
    """
    out = df.copy()
    out.columns = [c.strip() for c in out.columns]

    required = ["timestamp", "customer_phone"]
    missing = [c for c in required if c not in out.columns]
    if missing:
        raise ValueError(
            f"Missing required columns for downstream analysis: {missing}\n"
            "Your dataset must include these."
        )

    out["customer_phone"] = "cust_" + _hash_series(out["customer_phone"], salt=salt).str.slice(0, 16)

    if hash_interaction_id and "interaction_id" in out.columns:
        out["interaction_id"] = "int_" + _hash_series(out["interaction_id"], salt=salt).str.slice(0, 16)

    drop_cols: list[str] = []

    for c in ["account_number", "first_name", "last_name", "email", "address"]:
        if c in out.columns:
            drop_cols.append(c)

    # Free-text fields are PII magnets; drop by default
    if not keep_text:
        for c in ["rep_memo", "full_transcript"]:
            if c in out.columns:
                drop_cols.append(c)

    drop_cols = sorted(set(drop_cols))
    if drop_cols:
        out = out.drop(columns=drop_cols)

    return out, drop_cols


# ---------------------------
# Feature engineering + metrics
# ---------------------------

def _coerce_flag(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype="Int64")

    s = series.copy()

    if pd.api.types.is_numeric_dtype(s):
        return (s.fillna(0).astype(float) > 0).astype(int)

    s = s.astype(str).str.strip().str.lower()
    true_set = {"1", "true", "t", "yes", "y"}
    false_set = {"0", "false", "f", "no", "n", "nan", "none", ""}

    out = []
    for v in s:
        if v in true_set:
            out.append(1)
        elif v in false_set:
            out.append(0)
        else:
            try:
                out.append(1 if float(v) > 0 else 0)
            except Exception:
                out.append(0)

    return pd.Series(out, index=series.index, dtype="int64")


def detect_30dacc_column(df: pd.DataFrame) -> Optional[str]:
    cols = list(df.columns)
    preferred = ["flag_30DACC", "30DACC_flag", "flag30DACC", "dacc_30_flag", "dacc30_flag"]
    for c in preferred:
        if c in cols:
            return c

    for c in cols:
        cl = c.lower()
        if "30" in cl and ("dacc" in cl or "acc" in cl) and ("flag" in cl or "indicator" in cl):
            return c
    return None


def detect_crt_column(df: pd.DataFrame) -> Optional[str]:
    """
    Detect a CRT column. We assume CRT is numeric and typically in seconds.
    Accepts variants like CRT_seconds, CRT, crt, crt_sec, etc.
    """
    cols = list(df.columns)
    preferred = [
        "CRT_seconds", "CRT", "crt", "crt_seconds", "crt_sec", "CRT_sec", "CRTSeconds", "crtSeconds"
    ]
    for c in preferred:
        if c in cols:
            return c

    for c in cols:
        cl = c.lower()
        if "crt" in cl and ("sec" in cl or "seconds" in cl or cl == "crt"):
            return c

    for c in cols:
        if "crt" in c.lower():
            return c

    return None


def validate_columns(df: pd.DataFrame) -> None:
    required = ["timestamp", "customer_phone"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Missing required columns: {missing}\n"
            "Your dataset must include these for repeat-contact + time parsing."
        )


def clean_ids(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "rep_id" in out.columns:
        out["rep_id"] = out["rep_id"].astype(str).str.strip()
        out.loc[out["rep_id"].isin(["", "nan", "none", "None"]), "rep_id"] = "UNKNOWN"

    if "ivr_reason" in out.columns:
        out["ivr_reason"] = out["ivr_reason"].astype(str).str.strip()
        out.loc[out["ivr_reason"].isin(["", "nan", "none", "None"]), "ivr_reason"] = "UNKNOWN"

    return out


def scale_high_is_worse(x: pd.Series, low: float, high: float) -> pd.Series:
    """
    Scale a numeric series to 0..1 where higher is worse.
    - <= low -> 0
    - >= high -> 1
    - linear in between
    """
    s = pd.to_numeric(x, errors="coerce")
    out = (s - low) / (high - low)
    out = out.clip(lower=0.0, upper=1.0)
    return out.fillna(0.0)


def engineer_features(df: pd.DataFrame, repeat_window_hours: int = 72) -> pd.DataFrame:
    out = df.copy()

    out["ts"] = pd.to_datetime(out["timestamp"], errors="coerce", utc=False)

    for col in ["callback_flag", "cleanup_flag", "port_out_flag"]:
        if col in out.columns:
            out[col] = _coerce_flag(out[col])
        else:
            out[col] = 0

    dacc_col = detect_30dacc_column(out)
    if dacc_col is None:
        out["flag_30DACC"] = 0
    else:
        out["flag_30DACC"] = _coerce_flag(out[dacc_col])

    out = out.sort_values(["customer_phone", "ts"], kind="mergesort").reset_index(drop=True)

    window = pd.Timedelta(hours=repeat_window_hours)
    out["is_repeat_contact"] = 0
    out["repeat_count_in_window"] = 0

    for _, grp in out.groupby("customer_phone", sort=False):
        times = grp["ts"].values
        idxs = grp.index.values

        j = 0
        for i in range(len(times)):
            while j < i and pd.Timestamp(times[i]) - pd.Timestamp(times[j]) > window:
                j += 1
            prior_in_window = i - j
            out.loc[idxs[i], "repeat_count_in_window"] = prior_in_window
            out.loc[idxs[i], "is_repeat_contact"] = 1 if prior_in_window > 0 else 0

    def disposition_row(r) -> str:
        if r.get("port_out_flag", 0) == 1:
            return "Port-Out"
        if r.get("flag_30DACC", 0) == 1:
            return "Delayed Deactivation (30DACC)"
        if r.get("cleanup_flag", 0) == 1:
            return "Cleanup / Correction"
        if r.get("callback_flag", 0) == 1:
            return "Callback / Follow-up"
        return "Resolved / No Follow-up Observed"

    out["final_disposition"] = out.apply(disposition_row, axis=1)
    return out


def compute_trust_metrics(
    df: pd.DataFrame,
    crt_low: float = 480.0,
    crt_high: float = 900.0,
) -> pd.DataFrame:
    """
    Adds:
    - drift_score (includes CRT risk when present)
    - crt_value (normalized field name)
    - crt_risk (0-1 higher-worse, thresholds)
    """
    out = df.copy()

    for c in ["callback_flag", "cleanup_flag", "flag_30DACC", "port_out_flag", "is_repeat_contact"]:
        if c in out.columns:
            out[c] = _coerce_flag(out[c])

    crt_col = detect_crt_column(out)
    if crt_col is None:
        out["crt_value"] = np.nan
        out["crt_risk"] = 0.0
        has_crt = False
    else:
        out["crt_value"] = pd.to_numeric(out[crt_col], errors="coerce")
        out["crt_risk"] = scale_high_is_worse(out["crt_value"], low=crt_low, high=crt_high)
        has_crt = True

    # Drift score weights
    if has_crt:
        out["drift_score"] = (
            0.30 * out["flag_30DACC"]
            + 0.20 * out["cleanup_flag"]
            + 0.20 * out["port_out_flag"]
            + 0.15 * out["is_repeat_contact"]
            + 0.15 * out["crt_risk"]
        )
    else:
        out["drift_score"] = (
            0.35 * out["flag_30DACC"]
            + 0.25 * out["cleanup_flag"]
            + 0.20 * out["port_out_flag"]
            + 0.20 * out["is_repeat_contact"]
        )

    return out


# ---------------------------
# Aggregations
# ---------------------------

def summarize_by_rep(df: pd.DataFrame, crt_over: float = 900.0) -> pd.DataFrame:
    if "rep_id" not in df.columns:
        return pd.DataFrame()

    agg_map = {
        "calls": ("interaction_id", "count") if "interaction_id" in df.columns else ("ts", "count"),
        "cbr": ("callback_flag", "mean"),
        "cds": ("cleanup_flag", "mean"),
        "dacc30": ("flag_30DACC", "mean"),
        "por": ("port_out_flag", "mean"),
        "repeat_rate": ("is_repeat_contact", "mean"),
        "drift_score_avg": ("drift_score", "mean"),
    }

    has_crt = "crt_value" in df.columns and df["crt_value"].notna().any()
    if has_crt:
        dfx = df.copy()
        dfx["crt_over_flag"] = (pd.to_numeric(dfx["crt_value"], errors="coerce") > float(crt_over)).astype(int)
        agg_map.update({
            "crt_avg": ("crt_value", "mean"),
            "crt_p90": ("crt_value", lambda s: float(np.nanpercentile(pd.to_numeric(s, errors="coerce"), 90))),
            "crt_over_rate": ("crt_over_flag", "mean"),
            "crt_risk_avg": ("crt_risk", "mean"),
        })
        rep = dfx.groupby("rep_id", dropna=False).agg(**agg_map).reset_index()
    else:
        rep = df.groupby("rep_id", dropna=False).agg(**agg_map).reset_index()

    if has_crt and "crt_risk_avg" in rep.columns:
        rep["trust_score"] = (
            1.0
            - 0.25 * rep["dacc30"]
            - 0.20 * rep["cds"]
            - 0.25 * rep["por"]
            - 0.15 * rep["repeat_rate"]
            - 0.15 * rep["crt_risk_avg"]
        )
    else:
        rep["trust_score"] = (
            1.0
            - 0.30 * rep["dacc30"]
            - 0.25 * rep["cds"]
            - 0.25 * rep["por"]
            - 0.20 * rep["repeat_rate"]
        )

    rep = rep.sort_values(["trust_score", "calls"], ascending=[False, False]).reset_index(drop=True)
    return rep


def summarize_by_ivr(df: pd.DataFrame, crt_over: float = 900.0) -> pd.DataFrame:
    if "ivr_reason" not in df.columns:
        return pd.DataFrame()

    agg_map = {
        "calls": ("interaction_id", "count") if "interaction_id" in df.columns else ("ts", "count"),
        "cbr": ("callback_flag", "mean"),
        "cds": ("cleanup_flag", "mean"),
        "dacc30": ("flag_30DACC", "mean"),
        "por": ("port_out_flag", "mean"),
        "repeat_rate": ("is_repeat_contact", "mean"),
        "drift_score_avg": ("drift_score", "mean"),
    }

    has_crt = "crt_value" in df.columns and df["crt_value"].notna().any()
    if has_crt:
        dfx = df.copy()
        dfx["crt_over_flag"] = (pd.to_numeric(dfx["crt_value"], errors="coerce") > float(crt_over)).astype(int)
        agg_map.update({
            "crt_avg": ("crt_value", "mean"),
            "crt_p90": ("crt_value", lambda s: float(np.nanpercentile(pd.to_numeric(s, errors="coerce"), 90))),
            "crt_over_rate": ("crt_over_flag", "mean"),
            "crt_risk_avg": ("crt_risk", "mean"),
        })
        ivr = dfx.groupby("ivr_reason", dropna=False).agg(**agg_map).reset_index()
    else:
        ivr = df.groupby("ivr_reason", dropna=False).agg(**agg_map).reset_index()

    ivr = ivr.sort_values(["drift_score_avg", "calls"], ascending=[False, False]).reset_index(drop=True)
    return ivr


def summarize_repeat_customers(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby("customer_phone", dropna=False)
    cust = g.agg(
        calls=("interaction_id", "count") if "interaction_id" in df.columns else ("ts", "count"),
        any_port_out=("port_out_flag", "max"),
        any_30dacc=("flag_30DACC", "max"),
        any_cleanup=("cleanup_flag", "max"),
        any_callback=("callback_flag", "max"),
        avg_drift=("drift_score", "mean"),
        first_ts=("ts", "min"),
        last_ts=("ts", "max"),
    ).reset_index()

    cust = cust.sort_values(["calls", "avg_drift"], ascending=[False, False]).reset_index(drop=True)
    return cust


# ---------------------------
# Charts (basic + story + heatmaps)
# ---------------------------

def save_barh(df: pd.DataFrame, x: str, y: str, title: str, outpath: Path, top_n: int = 15) -> None:
    if df is None or df.empty or x not in df.columns or y not in df.columns:
        return
    d = df.head(top_n).copy().sort_values(x, ascending=True)

    plt.figure(figsize=(10, max(4, 0.35 * len(d) + 1)))
    plt.barh(d[y].astype(str), d[x].astype(float))
    plt.title(title)
    plt.xlabel(x)
    plt.tight_layout()
    plt.savefig(outpath, dpi=160)
    plt.close()


def save_hist(series: pd.Series, title: str, outpath: Path, bins: int = 20) -> None:
    if series is None or series.dropna().empty:
        return
    plt.figure(figsize=(10, 5))
    plt.hist(series.dropna().astype(float), bins=bins)
    plt.title(title)
    plt.xlabel(series.name or "value")
    plt.ylabel("count")
    plt.tight_layout()
    plt.savefig(outpath, dpi=160)
    plt.close()


def make_story_charts(rep: pd.DataFrame, ivr: pd.DataFrame, cust: pd.DataFrame, out_dir: Path) -> None:
    if ivr is not None and not ivr.empty and {"calls", "drift_score_avg"}.issubset(ivr.columns):
        plt.figure(figsize=(10, 6))
        plt.scatter(ivr["calls"], ivr["drift_score_avg"])
        plt.title("IVR hot zones: Drift vs Volume")
        plt.xlabel("Calls (volume)")
        plt.ylabel("Average drift_score")
        plt.tight_layout()
        plt.savefig(out_dir / "story_ivr_drift_vs_volume.png", dpi=160)
        plt.close()

    if rep is not None and not rep.empty and {"calls", "trust_score"}.issubset(rep.columns):
        plt.figure(figsize=(10, 6))
        plt.scatter(rep["calls"], rep["trust_score"])
        plt.title("Rep landscape: Trust Score vs Volume")
        plt.xlabel("Calls handled")
        plt.ylabel("Trust score (higher is better)")
        plt.tight_layout()
        plt.savefig(out_dir / "story_rep_trust_vs_volume.png", dpi=160)
        plt.close()

    if cust is not None and not cust.empty and "calls" in cust.columns:
        cust_sorted = cust.sort_values("calls", ascending=False).reset_index(drop=True)
        cust_sorted["cum_calls"] = cust_sorted["calls"].cumsum()
        cust_sorted["cum_calls_pct"] = cust_sorted["cum_calls"] / max(cust_sorted["calls"].sum(), 1)
        cust_sorted["cust_pct"] = (cust_sorted.index + 1) / max(len(cust_sorted), 1)

        plt.figure(figsize=(10, 6))
        plt.plot(cust_sorted["cust_pct"], cust_sorted["cum_calls_pct"])
        plt.title("Pareto: % of customers vs % of calls")
        plt.xlabel("Fraction of customers")
        plt.ylabel("Cumulative fraction of calls")
        plt.tight_layout()
        plt.savefig(out_dir / "story_customer_pareto_calls.png", dpi=160)
        plt.close()


def plot_corr_heatmap(corr: pd.DataFrame, title: str, outpath: Path) -> None:
    if corr is None or corr.empty:
        return

    cols = list(corr.columns)
    data = corr.values.astype(float)

    plt.figure(figsize=(max(8, 0.6 * len(cols) + 2), max(6, 0.6 * len(cols) + 2)))
    plt.imshow(data, aspect="auto")
    plt.title(title)
    plt.xticks(range(len(cols)), cols, rotation=45, ha="right")
    plt.yticks(range(len(cols)), cols)

    for i in range(len(cols)):
        for j in range(len(cols)):
            v = data[i, j]
            if np.isfinite(v):
                plt.text(j, i, f"{v:.2f}", ha="center", va="center", fontsize=8)

    plt.tight_layout()
    plt.savefig(outpath, dpi=160)
    plt.close()


def make_corr_heatmaps(rep_summary: pd.DataFrame, out_dir: Path) -> None:
    if rep_summary is None or rep_summary.empty:
        return

    numeric_cols = [
        c for c in rep_summary.columns
        if c != "rep_id" and pd.api.types.is_numeric_dtype(rep_summary[c])
    ]
    if len(numeric_cols) < 2:
        return

    signals = [c for c in numeric_cols if c != "trust_score"]
    if len(signals) >= 2:
        corr_signals = rep_summary[signals].corr(numeric_only=True)
        plot_corr_heatmap(
            corr_signals,
            title="Rep-level correlation heatmap (signals only)",
            outpath=out_dir / "rep_corr_heatmap_signals_only.png",
        )

    if "trust_score" in rep_summary.columns:
        corr_all = rep_summary[numeric_cols].corr(numeric_only=True)
        plot_corr_heatmap(
            corr_all,
            title="Rep-level correlation heatmap (including derived trust_score)",
            outpath=out_dir / "rep_corr_heatmap_with_trust.png",
        )


# ---------------------------
# Reporting
# ---------------------------

def _safe_div(n: float, d: float) -> float:
    return float(n) / float(d) if d else 0.0


def _drift_concentration_calls(df: pd.DataFrame, top_frac: float = 0.10) -> float:
    if len(df) == 0:
        return 0.0
    s = df["drift_score"].fillna(0).astype(float).sort_values(ascending=False)
    k = max(1, int(np.ceil(top_frac * len(s))))
    return _safe_div(s.head(k).sum(), s.sum())


def _drift_concentration_customers(df: pd.DataFrame, top_frac: float = 0.10) -> float:
    if len(df) == 0:
        return 0.0
    g = df.groupby("customer_phone", dropna=False).agg(
        calls=("interaction_id", "count") if "interaction_id" in df.columns else ("ts", "count"),
        avg_drift=("drift_score", "mean"),
    )
    g["total_drift_proxy"] = g["calls"].astype(float) * g["avg_drift"].astype(float)
    g = g.sort_values("total_drift_proxy", ascending=False)
    k = max(1, int(np.ceil(top_frac * len(g))))
    return _safe_div(g.head(k)["total_drift_proxy"].sum(), g["total_drift_proxy"].sum())


def write_summary_report(
    df: pd.DataFrame,
    rep_summary: pd.DataFrame,
    ivr_summary: pd.DataFrame,
    outpath: Path,
    min_calls_for_rank: int = 20,
    crt_low: float = 480.0,
    crt_high: float = 900.0,
    crt_over: float = 900.0,
) -> None:
    n = len(df)
    if n == 0:
        outpath.write_text("No rows found.\n", encoding="utf-8")
        return

    def pct(x: float) -> str:
        return f"{100*x:.1f}%"

    overall = {
        "CBR (callback rate)": df["callback_flag"].mean(),
        "CDS (cleanup rate)": df["cleanup_flag"].mean(),
        "30DACC rate": df["flag_30DACC"].mean(),
        "POR (port-out rate)": df["port_out_flag"].mean(),
        "Repeat contact rate": df["is_repeat_contact"].mean(),
        "Avg drift_score": df["drift_score"].mean(),
    }

    has_crt = "crt_value" in df.columns and df["crt_value"].notna().any()
    if has_crt:
        crt_vals = pd.to_numeric(df["crt_value"], errors="coerce")
        overall.update({
            "CRT avg (seconds)": float(np.nanmean(crt_vals)),
            "CRT p90 (seconds)": float(np.nanpercentile(crt_vals, 90)),
            f"CRT > {int(crt_over)}s rate": float(np.nanmean((crt_vals > crt_over).astype(float))),
            f"CRT risk avg (scaled {int(crt_low)}–{int(crt_high)}s)": float(np.nanmean(df["crt_risk"].astype(float))),
        })

    top10_calls_share = _drift_concentration_calls(df, top_frac=0.10)
    top10_cust_share = _drift_concentration_customers(df, top_frac=0.10)

    lines: list[str] = []
    lines.append("TRUST SIGNAL HEALTH — SUMMARY REPORT")
    lines.append("=" * 40)
    lines.append(f"Rows analyzed: {n}")
    lines.append("")

    lines.append("OVERALL RATES")
    for k, v in overall.items():
        if "Avg" in k or "CRT" in k or "p90" in k or "seconds" in k or "risk" in k:
            if "rate" in k.lower() and "CRT" in k:
                lines.append(f"- {k}: {pct(v)}")
            else:
                lines.append(f"- {k}: {v:.3f}" if "risk" in k else f"- {k}: {v:.1f}")
        else:
            lines.append(f"- {k}: {pct(v)}")
    lines.append("")

    if has_crt:
        lines.append("CRT INTERPRETATION")
        lines.append("- CRT is treated as a friction cost signal (higher is worse).")
        lines.append(f"- CRT risk is scaled: <= {int(crt_low)}s → 0, >= {int(crt_high)}s → 1.")
        lines.append("")

    lines.append("DRIFT CONCENTRATION (where risk clusters)")
    lines.append(f"- Top 10% of calls account for: {pct(top10_calls_share)} of total drift_score")
    lines.append(f"- Top 10% of customers account for: {pct(top10_cust_share)} of total drift_score (proxy)")
    lines.append("")

    lines.append("TOP 5 IVR REASONS BY DRIFT (avg)")
    if ivr_summary is None or ivr_summary.empty:
        lines.append("- (No ivr_reason column found; skipping IVR section)")
    else:
        base_cols = ["ivr_reason", "calls", "drift_score_avg", "por", "dacc30", "cds", "cbr"]
        extra_cols = ["crt_avg", "crt_p90", "crt_over_rate"] if "crt_avg" in ivr_summary.columns else []
        top_ivr = ivr_summary.head(5)[base_cols + extra_cols]
        for _, r in top_ivr.iterrows():
            msg = (
                f"- {r['ivr_reason']} | calls={int(r['calls'])} | drift={r['drift_score_avg']:.3f} "
                f"| POR={pct(r['por'])} | 30DACC={pct(r['dacc30'])} | CDS={pct(r['cds'])} | CBR={pct(r['cbr'])}"
            )
            if "crt_avg" in ivr_summary.columns:
                msg += f" | CRT_avg={r['crt_avg']:.0f}s | CRT_p90={r['crt_p90']:.0f}s | CRT>thr={pct(r['crt_over_rate'])}"
            lines.append(msg)
    lines.append("")

    if rep_summary is None or rep_summary.empty:
        lines.append("NO REP SUMMARY FOUND (missing rep_id?)")
        lines.append("")
        rep_rank = rep_summary
    else:
        qualified = rep_summary[rep_summary["calls"] >= min_calls_for_rank].copy()
        if len(qualified) >= 5:
            rep_rank = qualified
            used_threshold = True
        else:
            rep_rank = rep_summary.copy()
            used_threshold = False

        lines.append(f"REPS FOUND: {rep_summary['rep_id'].nunique(dropna=False)}")
        lines.append(f"REPS WITH >= {min_calls_for_rank} CALLS: {(rep_summary['calls'] >= min_calls_for_rank).sum()}")
        if not used_threshold:
            lines.append(
                f"NOTE: Fewer than 5 reps meet the >= {min_calls_for_rank} call threshold. "
                "Ranking uses all reps instead."
            )
        lines.append("")

    lines.append("TOP 5 REPS BY TRUST_SCORE (higher is better)")
    if rep_rank is None or rep_rank.empty:
        lines.append("- No reps available to rank.")
    else:
        cols = ["rep_id","calls","trust_score","drift_score_avg","por","dacc30","cds","cbr"]
        extra = ["crt_avg","crt_p90","crt_over_rate"] if "crt_avg" in rep_rank.columns else []
        top_rep = rep_rank.sort_values(["trust_score","calls"], ascending=[False, False]).head(5)[cols+extra]
        for _, r in top_rep.iterrows():
            msg = (
                f"- {r['rep_id']} | calls={int(r['calls'])} | trust_score={r['trust_score']:.3f} "
                f"| drift={r['drift_score_avg']:.3f} | POR={pct(r['por'])} | 30DACC={pct(r['dacc30'])} "
                f"| CDS={pct(r['cds'])} | CBR={pct(r['cbr'])}"
            )
            if "crt_avg" in rep_rank.columns:
                msg += f" | CRT_avg={r['crt_avg']:.0f}s | CRT_p90={r['crt_p90']:.0f}s | CRT>thr={pct(r['crt_over_rate'])}"
            lines.append(msg)

    lines.append("")
    lines.append("LOWEST 5 REPS BY TRUST_SCORE")
    if rep_rank is None or rep_rank.empty:
        lines.append("- No reps available to rank.")
    else:
        cols = ["rep_id","calls","trust_score","drift_score_avg","por","dacc30","cds","cbr"]
        extra = ["crt_avg","crt_p90","crt_over_rate"] if "crt_avg" in rep_rank.columns else []
        bottom_rep = rep_rank.sort_values(["trust_score","calls"], ascending=[True, False]).head(5)[cols+extra]
        for _, r in bottom_rep.iterrows():
            msg = (
                f"- {r['rep_id']} | calls={int(r['calls'])} | trust_score={r['trust_score']:.3f} "
                f"| drift={r['drift_score_avg']:.3f} | POR={pct(r['por'])} | 30DACC={pct(r['dacc30'])} "
                f"| CDS={pct(r['cds'])} | CBR={pct(r['cbr'])}"
            )
            if "crt_avg" in rep_rank.columns:
                msg += f" | CRT_avg={r['crt_avg']:.0f}s | CRT_p90={r['crt_p90']:.0f}s | CRT>thr={pct(r['crt_over_rate'])}"
            lines.append(msg)

    lines.append("")
    lines.append("BIGGEST COACHING OPPORTUNITIES (low trust_score + high volume)")
    if rep_rank is None or rep_rank.empty:
        lines.append("- No reps available to rank.")
    else:
        rep_rank2 = rep_rank.copy()
        rep_rank2["impact_risk"] = (1.0 - rep_rank2["trust_score"].astype(float)) * rep_rank2["calls"].astype(float)
        opp = rep_rank2.sort_values(["impact_risk"], ascending=[False]).head(5)
        for _, r in opp.iterrows():
            msg = (
                f"- {r['rep_id']} | calls={int(r['calls'])} | trust_score={r['trust_score']:.3f} "
                f"| impact_risk={r['impact_risk']:.1f} | drift={r['drift_score_avg']:.3f} "
                f"| POR={pct(r['por'])} | 30DACC={pct(r['dacc30'])} | CDS={pct(r['cds'])} | CBR={pct(r['cbr'])}"
            )
            if "crt_avg" in rep_rank2.columns:
                msg += f" | CRT_avg={r['crt_avg']:.0f}s | CRT_p90={r['crt_p90']:.0f}s | CRT>thr={pct(r['crt_over_rate'])}"
            lines.append(msg)

    lines.append("")
    lines.append("INTERPRETATION HINTS")
    lines.append("- Headline KPIs can look stable while drift indicators rise (30DACC/port-outs/repeats/CRT).")
    lines.append("- Cleanup isn’t failure; it’s often proof you’re fixing upstream mistakes (or catching bad inputs).")
    lines.append("- CRT is a cost signal: even if outcomes are stable, rising CRT suggests growing friction/complexity.")
    lines.append("- Concentration tells you whether to fix hotspots or address broad process issues.")

    outpath.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------
# Mathcheck workbook
# ---------------------------

def compute_framework_metrics(row: pd.Series, por_is_percent: bool = True) -> dict[str, float]:
    total_calls = float(row["TotalCalls"])
    callbacks = float(row["Callbacks"])
    cancels_per100 = float(row["CancelsPer100"])
    dacc_30 = float(row["DACC_30"])
    stay = float(row["StayConnected"])
    cleanup = float(row["CleanupCalls"])
    por = float(row["POR_percent"])

    cbr = 0.0 if total_calls == 0 else callbacks / total_calls
    cancels_norm = cancels_per100 / 100.0
    por_norm = (por / 100.0) if por_is_percent else por

    dfr = (30.0 * dacc_30) + cancels_norm
    dfr_star = dfr + por_norm

    cds = 0.0 if total_calls == 0 else cleanup / total_calls
    vrp = stay - (30.0 * dacc_30) - cancels_norm - por_norm

    tss_raw = (0.35 * vrp) + (0.25 * (1.0 - cds)) + (0.20 * (1.0 - dfr)) + (0.20 * (1.0 - por_norm))
    tss_clamped = max(0.0, min(1.0, tss_raw))

    return {
        "CBR": cbr,
        "CancelsNorm": cancels_norm,
        "POR_norm": por_norm,
        "DFR": dfr,
        "DFR_star": dfr_star,
        "CDS": cds,
        "VRP": vrp,
        "TSS_raw": tss_raw,
        "TSS_clamped": tss_clamped,
    }


def write_mathcheck_workbook(out_path: Path) -> Path:
    pseudo = pd.DataFrame([
        {"Scenario": "A High FCR, high 30DACC (delayed fallout)", "TotalCalls": 200, "Callbacks": 40, "CancelsPer100": 27.4, "DACC_30": 0.22, "StayConnected": 0.52, "CleanupCalls": 70, "POR_percent": 2.0},
        {"Scenario": "B Low FCR, improving stability",            "TotalCalls": 180, "Callbacks": 70, "CancelsPer100": 18.0, "DACC_30": 0.10, "StayConnected": 0.64, "CleanupCalls": 55, "POR_percent": 1.2},
        {"Scenario": "C Great FCR, silent exits",                 "TotalCalls": 220, "Callbacks": 30, "CancelsPer100": 12.0, "DACC_30": 0.08, "StayConnected": 0.70, "CleanupCalls": 40, "POR_percent": 6.5},
        {"Scenario": "D Heavy cleanup queue",                     "TotalCalls": 160, "Callbacks": 55, "CancelsPer100": 22.0, "DACC_30": 0.14, "StayConnected": 0.60, "CleanupCalls": 95, "POR_percent": 2.8},
        {"Scenario": "E Promo mismatch spike",                    "TotalCalls": 140, "Callbacks": 35, "CancelsPer100": 30.0, "DACC_30": 0.18, "StayConnected": 0.55, "CleanupCalls": 65, "POR_percent": 3.5},
        {"Scenario": "F Stable & clean",                          "TotalCalls": 250, "Callbacks": 45, "CancelsPer100": 10.0, "DACC_30": 0.05, "StayConnected": 0.78, "CleanupCalls": 35, "POR_percent": 1.0},
        {"Scenario": "G Volatile outcomes",                       "TotalCalls": 190, "Callbacks": 80, "CancelsPer100": 28.0, "DACC_30": 0.16, "StayConnected": 0.58, "CleanupCalls": 85, "POR_percent": 4.0},
        {"Scenario": "H Data edge case (0 calls)",                "TotalCalls": 0,   "Callbacks": 0,  "CancelsPer100": 0.0,  "DACC_30": 0.00, "StayConnected": 0.00, "CleanupCalls": 0,  "POR_percent": 0.0},
    ])

    py_out = pseudo.apply(lambda r: pd.Series(compute_framework_metrics(r)), axis=1)
    pseudo_with_py = pd.concat([pseudo, py_out.add_prefix("Py_")], axis=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "PseudoData"

    note = [
        "Trust Signal Health – Pseudo Metrics Workbook",
        "Inputs are fake (safe) example values to test the math.",
        "Assumptions: DACC_30 and StayConnected are proportions (e.g., 22.10% → 0.221). POR_percent is a percent (0–100).",
        "This sheet includes Excel formula outputs AND Python-computed outputs side-by-side, plus diff checks."
    ]
    for i, line in enumerate(note, start=1):
        ws.cell(row=i, column=1, value=line).font = Font(bold=True if i == 1 else False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    start_row = 6

    headers = [
        "Scenario","TotalCalls","Callbacks","CancelsPer100","DACC_30","StayConnected","CleanupCalls","POR_percent",
        "CBR","CancelsNorm","POR_norm","DFR","DFR_star","CDS","VRP","TSS_raw","TSS_clamped",
        "Py_CBR","Py_CancelsNorm","Py_POR_norm","Py_DFR","Py_DFR_star","Py_CDS","Py_VRP","Py_TSS_raw","Py_TSS_clamped",
        "Diff_CBR","Diff_CancelsNorm","Diff_POR_norm","Diff_DFR","Diff_DFR_star","Diff_CDS","Diff_VRP","Diff_TSS_raw","Diff_TSS_clamped",
        "PASS?"
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def ref(col, row_idx):
        return f"{get_column_letter(col)}{row_idx}"

    for i, r in enumerate(pseudo_with_py.itertuples(index=False), start=1):
        row_idx = start_row + i

        ws.cell(row=row_idx, column=1, value=r.Scenario)
        ws.cell(row=row_idx, column=2, value=r.TotalCalls)
        ws.cell(row=row_idx, column=3, value=r.Callbacks)
        ws.cell(row=row_idx, column=4, value=r.CancelsPer100)
        ws.cell(row=row_idx, column=5, value=r.DACC_30)
        ws.cell(row=row_idx, column=6, value=r.StayConnected)
        ws.cell(row=row_idx, column=7, value=r.CleanupCalls)
        ws.cell(row=row_idx, column=8, value=r.POR_percent)

        ws.cell(row=row_idx, column=9,  value=f"=IFERROR({ref(3,row_idx)}/{ref(2,row_idx)},0)")
        ws.cell(row=row_idx, column=10, value=f"={ref(4,row_idx)}/100")
        ws.cell(row=row_idx, column=11, value=f"={ref(8,row_idx)}/100")
        ws.cell(row=row_idx, column=12, value=f"=30*{ref(5,row_idx)}+{ref(10,row_idx)}")
        ws.cell(row=row_idx, column=13, value=f"={ref(12,row_idx)}+{ref(11,row_idx)}")
        ws.cell(row=row_idx, column=14, value=f"=IFERROR({ref(7,row_idx)}/{ref(2,row_idx)},0)")
        ws.cell(row=row_idx, column=15, value=f"={ref(6,row_idx)}-(30*{ref(5,row_idx)})-{ref(10,row_idx)}-{ref(11,row_idx)}")
        ws.cell(row=row_idx, column=16, value=f"=0.35*{ref(15,row_idx)}+0.25*(1-{ref(14,row_idx)})+0.20*(1-{ref(12,row_idx)})+0.20*(1-{ref(11,row_idx)})")
        ws.cell(row=row_idx, column=17, value=f"=MAX(0,MIN(1,{ref(16,row_idx)}))")

        py_vals = [r.Py_CBR, r.Py_CancelsNorm, r.Py_POR_norm, r.Py_DFR, r.Py_DFR_star, r.Py_CDS, r.Py_VRP, r.Py_TSS_raw, r.Py_TSS_clamped]
        for j, v in enumerate(py_vals):
            ws.cell(row=row_idx, column=18+j, value=float(v))

        for k in range(9):
            ws.cell(row=row_idx, column=27+k, value=f"=ABS({ref(9+k,row_idx)}-{ref(18+k,row_idx)})")

        ws.cell(row=row_idx, column=36, value=f"=IF(MAX({ref(27,row_idx)}:{ref(35,row_idx)})<1E-9,\"PASS\",\"CHECK\")")

        for col in range(1, 37):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = ws["A7"]

    table_ref = f"A{start_row}:AJ{start_row + len(pseudo_with_py)}"
    tab = Table(displayName="TrustSignalPseudo", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ---------------------------
# Subcommands
# ---------------------------

def cmd_run(args: argparse.Namespace) -> int:
    repo_root = find_repo_root(Path.cwd())
    data_dir = repo_root / "data"
    out_dir = repo_root / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    in_path = get_input_file(data_dir, args.file)
    df_raw = load_table(in_path)

    if args.scrub_pii:
        df_work, dropped = scrub_pii_df(
            df_raw,
            salt=args.salt,
            hash_interaction_id=True,
            keep_text=args.keep_text,
        )
        scrubbed_out = out_dir / "calls_scrubbed.csv"
        df_work.to_csv(scrubbed_out, index=False)
    else:
        df_work = df_raw
        dropped = []
        scrubbed_out = None

    df_work = clean_ids(df_work)
    validate_columns(df_work)

    df2 = engineer_features(df_work, repeat_window_hours=args.repeat_window_hours)
    df2 = compute_trust_metrics(df2, crt_low=args.crt_low, crt_high=args.crt_high)

    calls_scored_path = out_dir / "calls_scored_cleaned.csv"
    df2.to_csv(calls_scored_path, index=False)

    rep_summary = summarize_by_rep(df2, crt_over=args.crt_over)
    ivr_summary = summarize_by_ivr(df2, crt_over=args.crt_over)
    cust_summary = summarize_repeat_customers(df2)

    rep_path = out_dir / "rep_summary.csv"
    ivr_path = out_dir / "ivr_summary.csv"
    cust_path = out_dir / "customer_repeat_summary.csv"

    rep_summary.to_csv(rep_path, index=False)
    ivr_summary.to_csv(ivr_path, index=False)
    cust_summary.to_csv(cust_path, index=False)

    report_path = out_dir / "summary_report.txt"
    write_summary_report(
        df2,
        rep_summary=rep_summary,
        ivr_summary=ivr_summary,
        outpath=report_path,
        min_calls_for_rank=args.min_calls_for_rank,
        crt_low=args.crt_low,
        crt_high=args.crt_high,
        crt_over=args.crt_over,
    )

    save_barh(
        rep_summary,
        x="trust_score",
        y="rep_id",
        title="Top reps by trust_score (higher is better)",
        outpath=out_dir / "rep_trust_score_top.png",
        top_n=15,
    )
    save_barh(
        ivr_summary,
        x="drift_score_avg",
        y="ivr_reason",
        title="IVR reasons with highest average drift_score",
        outpath=out_dir / "ivr_drift_top.png",
        top_n=15,
    )
    save_hist(
        df2["drift_score"],
        title="Distribution of drift_score (row-level)",
        outpath=out_dir / "drift_score_hist.png",
        bins=20,
    )

    if "crt_value" in df2.columns and df2["crt_value"].notna().any():
        save_hist(
            df2["crt_value"],
            title="Distribution of CRT (seconds)",
            outpath=out_dir / "crt_seconds_hist.png",
            bins=25,
        )

    make_story_charts(rep_summary, ivr_summary, cust_summary, out_dir)
    make_corr_heatmaps(rep_summary, out_dir)

    if args.with_mathcheck:
        math_path = write_mathcheck_workbook(out_dir / "Trust_Signal_Health_Psuedo_Math_Check.xlsx")
    else:
        math_path = None

    print(f"Loaded:  {in_path}")
    if scrubbed_out is not None:
        print(f"Wrote:   {scrubbed_out}")
        if dropped:
            print(f"Dropped columns (PII/text): {dropped}")
    print(f"Wrote:   {calls_scored_path}")
    print(f"Wrote:   {rep_path}")
    print(f"Wrote:   {ivr_path}")
    print(f"Wrote:   {cust_path}")
    print(f"Wrote:   {report_path}")
    print("Wrote:   charts (*.png) into /output")
    if math_path is not None:
        print(f"Wrote:   {math_path}")
    print("Done.")
    return 0


def cmd_scrub(args: argparse.Namespace) -> int:
    repo_root = find_repo_root(Path.cwd())
    data_dir = repo_root / "data"
    out_dir = repo_root / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    in_path = get_input_file(data_dir, args.file)
    df_raw = load_table(in_path)

    df_scrub, dropped = scrub_pii_df(
        df_raw,
        salt=args.salt,
        hash_interaction_id=True,
        keep_text=args.keep_text,
    )

    out_path = out_dir / "calls_scrubbed.csv"
    df_scrub.to_csv(out_path, index=False)

    print(f"Loaded: {in_path}")
    print(f"Wrote:  {out_path}")
    if dropped:
        print(f"Dropped columns: {dropped}")
    else:
        print("Dropped columns: (none)")
    print("Done.")
    return 0


def cmd_charts(args: argparse.Namespace) -> int:
    repo_root = find_repo_root(Path.cwd())
    out_dir = repo_root / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    rep_path = out_dir / "rep_summary.csv"
    ivr_path = out_dir / "ivr_summary.csv"
    cust_path = out_dir / "customer_repeat_summary.csv"

    if not rep_path.exists() or not ivr_path.exists() or not cust_path.exists():
        raise FileNotFoundError(
            "Missing one or more summary files in /output.\n"
            "Run: python src/analyze_trust_signals.py run\n"
            "Then rerun charts."
        )

    rep = pd.read_csv(rep_path)
    ivr = pd.read_csv(ivr_path)
    cust = pd.read_csv(cust_path)

    make_story_charts(rep, ivr, cust, out_dir)
    make_corr_heatmaps(rep, out_dir)

    print("Wrote story charts + correlation heatmaps into /output")
    return 0


def cmd_mathcheck(args: argparse.Namespace) -> int:
    repo_root = find_repo_root(Path.cwd())
    out_dir = repo_root / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / "Trust_Signal_Health_Psuedo_Math_Check.xlsx"
    saved = write_mathcheck_workbook(out_path)
    print(f"Wrote: {saved}")
    return 0


# ---------------------------
# CLI
# ---------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Trust Signal analysis (single-script pipeline).")
    sub = parser.add_subparsers(dest="cmd")

    p_run = sub.add_parser("run", help="Scrub (optional), analyze, report, and generate charts.")
    p_run.add_argument(
        "--file",
        default="novawireless_synthetic_calls.csv",
        help="Input filename inside /data (default: novawireless_synthetic_calls.csv).",
    )
    p_run.add_argument("--repeat_window_hours", type=int, default=72)
    p_run.add_argument("--min_calls_for_rank", type=int, default=20)

    p_run.add_argument("--crt_low", type=float, default=480.0, help="CRT low threshold (seconds) for risk scaling.")
    p_run.add_argument("--crt_high", type=float, default=900.0, help="CRT high threshold (seconds) for risk scaling.")
    p_run.add_argument("--crt_over", type=float, default=900.0, help="CRT 'over threshold' rate cutoff (seconds).")

    p_run.add_argument(
        "--no_scrub_pii",
        dest="scrub_pii",
        action="store_false",
        help="Disable PII scrubbing.",
    )
    p_run.set_defaults(scrub_pii=True)

    p_run.add_argument(
        "--salt",
        default="trust-signals-salt",
        help="Salt for hashing identifiers (keep constant for stable IDs).",
    )
    p_run.add_argument(
        "--keep_text",
        action="store_true",
        help="Keep free-text fields (NOT recommended for GitHub).",
    )
    p_run.add_argument(
        "--with_mathcheck",
        action="store_true",
        help="Also generate the pseudo math validation workbook in /output.",
    )
    p_run.set_defaults(func=cmd_run)

    p_scrub = sub.add_parser("scrub", help="Only scrub PII and write output/calls_scrubbed.csv")
    p_scrub.add_argument(
        "--file",
        default="novawireless_synthetic_calls.csv",
        help="Input filename inside /data (default: novawireless_synthetic_calls.csv).",
    )
    p_scrub.add_argument("--salt", default="trust-signals-salt")
    p_scrub.add_argument("--keep_text", action="store_true")
    p_scrub.set_defaults(func=cmd_scrub)

    p_charts = sub.add_parser("charts", help="Regenerate story charts + correlation heatmaps from /output summaries")
    p_charts.set_defaults(func=cmd_charts)

    p_math = sub.add_parser("mathcheck", help="Generate pseudo math validation workbook in /output")
    p_math.set_defaults(func=cmd_mathcheck)

    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_parser()
    argv2 = argv if argv is not None else (sys.argv[1:] or ["run"])
    args = parser.parse_args(argv2)

    # If user runs "python src/analyze_trust_signals.py" with no args,
    # argv2 becomes ["run"], so args.func exists.
    if not hasattr(args, "func"):
        parser.print_help()
        return 2

    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
